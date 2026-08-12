import pathlib, re, unicodedata
from difflib import SequenceMatcher
import pdfplumber
from openpyxl import load_workbook
import csv

pdf = next(pathlib.Path("dokumanlar/mufredat").glob("deneme_*.pdf"))
wb = load_workbook("dokumanlar/ogrenci_listesi.xlsx", read_only=True, data_only=True)
students = []
for sinif, no, ad in list(wb.active.iter_rows(min_row=2, values_only=True)):
    if no is not None and ad:
        students.append({"sinif": str(sinif), "no": str(int(no)), "ad": str(ad)})

def norm(s):
    s = s.upper().translate(str.maketrans("ÇĞİÖŞÜI", "CGIOSUI"))
    s = unicodedata.normalize("NFKD", s)
    return " ".join(re.sub(r"[^A-Z0-9 ]", " ", s).split())

for s in students:
    s["norm"] = norm(s["ad"])

rows = []
row_re = re.compile(r"^(\d+)\s+(\d+)\s+(.+?)\s+(10-(?:A|B|C|Ç|D|XX))\s*(.*)$")
with pdfplumber.open(pdf) as doc:
    for page_no in (0, 1):
        for line in (doc.pages[page_no].extract_text() or "").splitlines():
            m = row_re.match(line.strip())
            if not m:
                continue
            rank, no, name, old_class, rest = m.groups()
            nums = rest.split()
            if len(nums) < 12:
                continue
            try:
                values = []
                for i in range(0, 12, 3):
                    values.append((int(nums[i]), int(nums[i+1]), nums[i+2]))
            except ValueError:
                continue
            rows.append({"rank": int(rank), "no": no, "name": name, "old_class": old_class, "values": values})

by_no = {s["no"]: s for s in students}
for r in rows:
    rnorm = norm(r["name"])
    exact_no = by_no.get(r["no"]) if r["no"] != "0" else None
    if exact_no:
        r["match"] = exact_no
        r["method"] = "okul numarası"
        r["score"] = SequenceMatcher(None, rnorm, exact_no["norm"]).ratio()
    else:
        old_branch = r["old_class"].split("-", 1)[1]
        candidates = []
        for s in students:
            branch_bonus = 0.18 if s["sinif"].endswith("-" + old_branch) else 0
            a, b = set(rnorm.split()), set(s["norm"].split())
            overlap = len(a & b) / max(1, min(len(a), len(b)))
            seq = SequenceMatcher(None, rnorm, s["norm"]).ratio()
            candidates.append((0.55 * overlap + 0.27 * seq + branch_bonus, s, overlap, seq))
        candidates.sort(key=lambda x: x[0], reverse=True)
        best = candidates[0]
        r["candidate"] = best[1]
        r["score"] = best[0]
        if best[0] >= 0.62:
            r["match"] = best[1]
            r["method"] = "isim parçaları + şube"
        else:
            r["match"] = None
            r["method"] = "eşleşmedi"

# Okul numarasının değiştiği/yeniden kullanıldığı ve deneme dosyasında takma ad
# bulunan özel kayıtlar, ad + şube birlikte değerlendirilerek düzeltilir.
manual = {
    1: ("59", "AKİF ve PAK ortak; 10-A -> 11-A"),
    3: (None, "BOJACK için mevcut listede güvenilir ortak nokta yok"),
    33: ("193", "NİLDA KARADAŞ birebir; 10-D -> 11-D, eski numara 195"),
    49: ("194", "EFE ortak ve 10-B -> 11-B; tek isim nedeniyle orta güven"),
    50: (None, "LİONEL MESSİ adı ve 10-D şubesi mevcut 10 numarayla uyuşmuyor"),
    54: (None, "AHMET tek başına ve 10-XX sınıfı ayırt edici değil"),
    63: ("393", "İLKER MURAT KURT birebir; 10-D -> 11-D, eski numara 384"),
    66: (None, "MERT tek başına; aynı şubede güvenilir aday yok"),
    71: (None, "AHMET MEHMET ve 10-Ç için güvenilir karşılık yok"),
    72: ("4646", "HALİL ortak ve 10-C -> 11-C; tek isim nedeniyle orta güven"),
}
for r in rows:
    if r["rank"] in manual:
        no, reason = manual[r["rank"]]
        r["match"] = by_no.get(no) if no else None
        r["method"] = "manuel ortak nokta" if no else "aktarılmadı"
        r["reason"] = reason
    elif r.get("match"):
        same_no = r["no"] == r["match"]["no"] and r["no"] != "0"
        same_name = norm(r["name"]) == r["match"]["norm"]
        same_branch = r["old_class"].split("-", 1)[1] == r["match"]["sinif"].split("-", 1)[1]
        common = sorted(set(norm(r["name"]).split()) & set(r["match"]["norm"].split()))
        r["reason"] = f"{'numara aynı' if same_no else 'numara farklı'}; {'ad aynı' if same_name else 'ortak ad parçaları: ' + ', '.join(common)}; {'şube aynı' if same_branch else 'şube farklı'}"
    else:
        r["reason"] = "Güvenilir eşleşme bulunamadı"

print("rows", len(rows), "students", len(students))
def birebir(r):
    return bool(r.get("match") and r["no"] != "0" and r["no"] == r["match"]["no"] and norm(r["name"]) == r["match"]["norm"] and r["old_class"].split("-",1)[1] == r["match"]["sinif"].split("-",1)[1])
exact = [r for r in rows if birebir(r)]
nonexact = [r for r in rows if not birebir(r)]
print("exact_name", len(exact), "nonexact_or_unmatched", len(nonexact))
for r in nonexact:
    m = r.get("match")
    cand = r.get("candidate")
    target = m or cand
    target_text = f"{target['no']} {target['ad']} {target['sinif']}" if target else "YOK"
    print(f"{r['rank']:02d}|pdf_no={r['no']}|pdf={r['name']}|{r['old_class']}|target={target_text}|method={r['method']}|score={r['score']:.3f}")

out = pathlib.Path("output")
out.mkdir(exist_ok=True)
with (out / "deneme_eslestirmeleri.csv").open("w", newline="", encoding="utf-8-sig") as f:
    w = csv.writer(f)
    w.writerow(["sira", "pdf_okul_no", "pdf_adi", "gecen_yil_sinifi", "mevcut_okul_no", "mevcut_adi", "mevcut_sinifi", "durum", "gerekce", "turkce_d", "turkce_y", "sosyal_d", "sosyal_y", "matematik_d", "matematik_y", "fen_d", "fen_y"])
    for r in rows:
        m = r.get("match")
        durum = "birebir" if birebir(r) else ("ortak_noktayla_eslesti" if m else "aktarilmadi")
        flat = [x for pair in r["values"] for x in pair[:2]]
        w.writerow([r["rank"], r["no"], r["name"], r["old_class"], m["no"] if m else "", m["ad"] if m else "", m["sinif"] if m else "", durum, r["reason"], *flat])

with (out / "deneme_eslestirme_raporu.md").open("w", encoding="utf-8") as f:
    f.write("# Geçen Yıl Denemesi - Öğrenci Eşleştirme Raporu\n\n")
    f.write("Sınav: **10. Sınıf Maarif Süreç Değerlendirme**  \n")
    f.write("Kaynak: `dokumanlar/mufredat/deneme_geçmiş.pdf`  \n")
    f.write(f"Toplam sonuç: **{len(rows)}** | Birebir: **{len(exact)}** | Ortak noktayla eşleşen: **{sum(bool(r.get('match')) and not birebir(r) for r in rows)}** | Aktarılmayan: **{sum(not r.get('match') for r in rows)}**\n\n")
    f.write("## Birebir tutmayan kayıtların tek tek değerlendirmesi\n\n")
    for i, r in enumerate(nonexact, 1):
        m = r.get("match")
        hedef = f"{m['no']} - {m['ad']} ({m['sinif']})" if m else "Aktarılmadı"
        f.write(f"{i}. **PDF:** {r['no']} - {r['name']} ({r['old_class']})  \n   **Sonuç:** {hedef}  \n   **Gerekçe:** {r['reason']}\n\n")
